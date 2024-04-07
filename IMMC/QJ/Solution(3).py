# Algorithm Part 1:
# Objective and realizing methods
# defining the justice index, J, as the total value gained divided by total value expected.
# when J isn't the save, give the cargo to the one with lowest J
# when there're same Js, give the cargo to the one with highest V
# 目的及实现方法
# 将正义指数 J 定义为获得的总价值除以预期的总价值。
# 当J不是保存时，将货物交给J最低的那个
# 当J相同时，将货物交给V最高的那个
import math
from itertools import repeat
import IMMCMichPack as mp
import numpy
import pandas as pd
from openpyxl import load_workbook
import string
####################################################### Division line

num_of_rows = 30
num_of_columns = 5
V_size = [num_of_columns,num_of_rows]

# Algorithm Part 2:
# Modification 1: Group two people together, calculate their J as a whole
# Modification 2: Based on an additional matrix, formulate and update the value matrix by the distribution result. Note that in this scenario the T_Expected is no longer constant.

# Steps for m2: {generate V-> [distribute cargo -> update V -> update T -> update J] -> distribute cargo} loop 


###### DATA loading
CorrelationWorkBook = load_workbook("/Volumes/Personal/Coding/Python/Main/IMMC/ItermCorrelation.xlsx")  # Work Book
Gain_WorkSheet = CorrelationWorkBook.get_sheet_by_name('Gain-Table')  # Work Sheet
Loss_WorkSheet = CorrelationWorkBook.get_sheet_by_name('Loss-Table')  # Work Sheet 

Correlation_Column_Names = list(string.ascii_uppercase[1:])
Correlation_Column_Names.append('AA')
Correlation_Column_Names.append('AB')
Correlation_Column_Names.append('AC')
Correlation_Column_Names.append('AD')
Correlation_Column_Names.append('AE')


Gain_Matrix = []
Loss_Matrix = []

for i in Correlation_Column_Names:
  column = Gain_WorkSheet[i]  # Column
  column_list = [column[x].value for x in range(1,1+V_size[1])] # take the number datas
  Gain_Matrix.append(column_list) # The V matrix, by name and item number

for i in Correlation_Column_Names:
  column = Loss_WorkSheet[i]  # Column
  column_list = [column[x].value for x in range(1,1+V_size[1])] # take the number datas
  Loss_Matrix.append(column_list) # The V matrix, by name and item number







Correlation_Coefficient = 5 #SASASASASASASASASASA





wb = load_workbook("/Volumes/Personal/Coding/Python/Main/IMMC/RawData 2.xlsx")  # Work Book
ws = wb.get_sheet_by_name('Ordered-Before Transition')  # Work Sheet



# Average of a list
def Expectation(Vsum):
  sum = 0
  for i in Vsum:
    sum+=i
  return sum/len(Vsum)
# Standard variance
def StandardVariance(Vsum):
  V = 0
  E = Expectation(Vsum)
  for i in Vsum:
    J = i - E 
    K = J**2
    V +=K
  S = math.sqrt(V/len(Vsum))
  return S
# Absolute Variance
def AbsoluteVariance(Vsum):
  V = 0
  E = Expectation(Vsum)
  for i in Vsum:
    J = i - E 
    K = math.sqrt(J**2)
    V += K
  S = V/len(Vsum)
  return S




Value_Matrix_Column_Names=list(string.ascii_uppercase[1:6]) # B to AE

def Read_Matrix(Sheet,ColumnRange,RowRange,Matrix_Name):
  for i in ColumnRange:
    column = Sheet[i]  # Column
    column_list = [column[x].value for x in RowRange] # take the number datas
    Matrix_Name.append(column_list) # The V matrix, by name and item number
  return None

Value_Matrix_Untransposed = []
Read_Matrix(ws,Value_Matrix_Column_Names,range(1,1+V_size[1]),Value_Matrix_Untransposed)

Distribution_Order = []
Read_Matrix(ws,['A'],range(1,V_size[1]+1),Distribution_Order)
Distribution_Order = Distribution_Order[0]

# Transpose the matrix

InterMediate = zip(*Value_Matrix_Untransposed)

Value_Matrix_Initial=[list(x) for x in InterMediate]


####### Check OK
# V= [column][row]

Vsum=[0 for i in range(5)]

def Matrix_ixj(i,j):
  Matri = [[0 for col in range(j)] for row in range(i)]
  return Matri

def Calc_VD_Matrix(Value_Matrix,Distribution_Matrix): # D,V
  VD_Matrix = Matrix_ixj(num_of_rows,num_of_columns)  
  for rows in range(num_of_rows):
    for cols in range(num_of_columns):
      VD_Matrix[rows][cols] = Value_Matrix[rows][cols]*Distribution_Matrix[rows][cols]
  return VD_Matrix


def Calc_T_Expected(Value_Matrix):
  T_Expected = [] #  not Constant of the expected total value
  for col in range(num_of_columns):
    S = 0
    for row in range(num_of_rows):
      S += Value_Matrix[row][col]
    T_Expected.append(S)
  return T_Expected

def Justice_Value(col,T_Gained,T_Expected):
  r = T_Gained[col]/T_Expected[col]
  return r



T_Gained = [0 for i in range(num_of_columns)] # Variable changes with the distribution of cargos
T_Expected = [0 for i in range(num_of_columns)]
J = [0 for i in range(num_of_columns)]
Distribution_Matrix = [[0,0,0,0,0] for i in range(num_of_rows)]




Value_Matrix = Value_Matrix_Initial

# Gain_Matrix[row][col] = the change in value of [col] after the GAIN of [row]
# Loss_Matrix[row][col] = the change in value of [col] after the LOSS of [row]








def SubMatrix(Matrix,column_list):
  R=[]
  for row in range(len(Matrix)):
    R.append([])
    for col in column_list:
      R[row].append(Matrix[row][col])
  return R

def J_for_cargo(row,col,Matrix):
  T_Expected = Calc_T_Expected(Matrix)
  Jc = Matrix[row][col] / T_Expected
  return Jc

def Update_Value_Matrix(item_distributed,who_to_give): #item_distributed is the location or index of the itgem in the order list
  V_Current = Value_Matrix 
  who_doesnot_get = [0,1,2,3,4]
  who_doesnot_get.pop(who_doesnot_get.index(who_to_give))
  # Change only after the cargo
  # Range_to_be_changed = range(item_distributed+1,len(Distribution_Order))
  # Change everything
  Range_to_be_changed = range(0,len(Distribution_Order)) 
  for changing_position in Range_to_be_changed: ##### If you wanna go from the beginning, it is really great
    item_order = Distribution_Order[item_distributed]
    changing_item_order = Distribution_Order[changing_position]
    # Q = Gain_Matrix[item_order-1][changing_item_order-1]
    # print(Q)
    R = (100+Correlation_Coefficient*Gain_Matrix[item_order][changing_item_order])/100*V_Current[changing_position][who_to_give] # Both the variables are item number not location 
    
    Value_Matrix[changing_position][who_to_give] = R
    for else_people in who_doesnot_get:
      R = (100+Correlation_Coefficient*Loss_Matrix[item_order][changing_item_order])/100*V_Current[changing_position][else_people] # Both the variables are item number not location n
      Value_Matrix[changing_position][else_people] = R
  return None
# This is right!!!

# Please notice that the numbering in the ItermCorrelation Matrix and the RawData Matrix is not the same.

current_row_number = 0
while current_row_number<num_of_rows:
  T_Expected = Calc_T_Expected(Value_Matrix)

  for col in range(num_of_columns):
    J[col]=Justice_Value(col,T_Gained,T_Expected)

  # Need a function that checks whether there're any repeated values in the list.
  whether_repeat_values = any(J.count(x) > 1 for x in J)
  
  # print(whether_repeat_values)
  
  if whether_repeat_values: 
    cut_list = []
    for i in range(len(J)):
      if J[i]==0:
        cut_list.append(i)
    sub_V=SubMatrix(Value_Matrix,cut_list)

    who_to_give = Value_Matrix[current_row_number].index(max(sub_V[current_row_number]))
  else:
    who_to_give = J.index(min(J)) # Anchor point. Via changing this, the optimization instance variable can be changed.
  Update_Value_Matrix(current_row_number,who_to_give)
  Distribution_Matrix[current_row_number][who_to_give] = 1 # Record the resultant matrix
  T_Gained[who_to_give]+=Value_Matrix[current_row_number][who_to_give]
  current_row_number += 1




VDM = Calc_VD_Matrix(Value_Matrix,Distribution_Matrix)


print(pd.DataFrame(VDM))
print(J)
# breakpoint()

rawData = pd.DataFrame(Value_Matrix)
newRawData = pd.DataFrame({'0': rawData.iloc[:, 0], '1': rawData.iloc[:, 1]})
sortedList = pd.DataFrame(VDM)
avgSortedList = sortedList.replace(0,numpy.NaN)
newSortList = pd.DataFrame({'0': sortedList.iloc[:, 0], '1': sortedList.iloc[:, 1]})
newAvgList = newSortList.replace(0,numpy.NaN)
avgSortedListToBePrinted = pd.DataFrame({'0': avgSortedList.sum()}).transpose()
adList = []
adList.append(mp.getad(avgSortedList))

newMax = []
newMaxId = []
mask = []

# Calc Final T Excepted
New_T_Excepted = rawData.iloc[:,0:3].sum()
New_T_Gained = sortedList.sum()
New_Justice_Value = New_T_Gained/New_T_Excepted

print(New_T_Gained)
print(New_Justice_Value)
# breakpoint()


for j in range(30):
    mask.append(j) if newSortList.iloc[j, 0].item() != 0 or newSortList.iloc[j, 1].item() != 0 else None


iterationCount = 0
for j in mask:
    newMax.append(newRawData.iloc[j].max())
    newMaxId.append(int(newRawData.iloc[j].idxmax()))
    newSortList.loc[j] = list(repeat(0, newMaxId[iterationCount])) + list(repeat(newMax[iterationCount], 1)) + list(
        repeat(0, (1 - newMaxId[iterationCount])))
    newAvgList.loc[j] = list(repeat(numpy.NaN, newMaxId[iterationCount])) + list(repeat(newMax[iterationCount], 1)) + list(
        repeat(numpy.NaN, (1 - newMaxId[iterationCount])))
    iterationCount += 1

for r in range(2):
    sortedList.iloc[:, r] = newSortList.iloc[:, r]
    avgSortedList.iloc[:, r] = newSortList.iloc[:, r]

jDataFrame = pd.Series({"Value of J": J})
printList = sortedList._append(avgSortedListToBePrinted._append(avgSortedList.sum().transpose(), ignore_index=True)
                               , ignore_index=True)
printList = printList._append(jDataFrame, ignore_index=True)
# printList.to_excel("NewDataOut_GPL.xlsx",sheet_name="AB+C")
with pd.ExcelWriter("NewDataOut_GPL.xlsx", mode="a", if_sheet_exists="replace") as writer:
    printList.to_excel(writer, sheet_name="AB+C+D+E")
mp.printdata(sortedList, avgSortedList, adList)
















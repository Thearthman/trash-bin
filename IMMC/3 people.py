import Settings
import Distribution as d
# Algorithm Part 1:
# Objective and realizing methods
# defining the justice index, J, as the total value gained divided by total value expected.
# when J isn't the save, give the cargo to the one with lowest J
# when there're same Js, give the cargo to the one with highest V
# 目的及实现方法
# 将正义指数 J 定义为获得的总价值除以预期的总价值。
# 当J不是保存时，将货物交给J最低的那个
# 当J相同时，将货物交给V最高的那个



from itertools import repeat
import matplotlib as plt
import numpy as np
import math
from openpyxl import load_workbook



def Take_Selected_Columns(Cooperation_Group):
  # COLUMN
  colums_taken = []
  for item in Cooperation_Group:
    if item.__class__ == list:
      for subitem in item:
        colums_taken.append(subitem)
    else:
      colums_taken.append(item)
  # NEW V MATRIX
  return(colums_taken)



num_of_rows = 30
num_of_columns = 3



V_size = [num_of_columns,num_of_rows]
import string
### Output the data to xlsx table.:
import xlsxwriter as xw
output_table =  xw.Workbook('/Users/Disk/Programs/IMMC2024winter/Code/tables/output_table.xlsx')
wb = load_workbook("/Users/Disk/Programs/IMMC2024winter/Code/tables/RawData.xlsx")  # Work Book
ws = wb.get_sheet_by_name('Ordered-Before Transition')  # Work Sheet


def Read_Matrix(Sheet,ColumnRange,RowRange,Matrix_Name):
  for i in ColumnRange:
    column = Sheet[i]  # Column
    column_list = [column[x].value for x in RowRange] # take the number datas
    Matrix_Name.append(column_list) # The V matrix, by name and item number
  return None

Distribution_Order = []
Read_Matrix(ws,['A'],range(1,31),Distribution_Order)
Distribution_Order = Distribution_Order[0]





def PRINT_Matrix(Sheet,Matrix):
  # Sheet = output_table.add_worksheet(SheetName)
  for rows in range(len(Matrix)):
    for cols in range(len(Matrix[rows])):
      Sheet.write(rows,cols,Matrix[rows][cols]) 
  return None

def PRINT_Verticle_Column(Sheet,vec,column):
  #if whether_add_new:
  #  SheetName = output_table.add_worksheet(SheetName)
  for rows in range(len(vec)):
    Sheet.write(rows,column,vec[rows])  
  return None

def PRINT_Horizontal_Row(Sheet,vec,row):
  for columns in range(len(vec)):
    Sheet.write(row,columns,vec[columns])  
  return None  



# Average of a list
def Expectation(Vsum):
  sum = 0
  for i in Vsum:
    sum+=i
  return sum/len(Vsum)
# Standard variance
def StandardVariance(Vsum):
  V = 0
  E = Expectation(Vsum)
  for i in Vsum:
    J = i - E 
    K = J**2
    V +=K
  S = math.sqrt(V/len(Vsum))
  return S
# Absolute Variance
def AbsoluteVariance(Vsum):
  V = 0
  E = Expectation(Vsum)
  for i in Vsum:
    J = i - E 
    K = math.sqrt(J**2)
    V += K
  S = V/len(Vsum)
  return S



Value_Matrix_Column_Names=list(string.ascii_uppercase[1:6]) # B to AE


Value_Matrix_Untransposed = []
Read_Matrix(ws,Value_Matrix_Column_Names,range(1,1+V_size[1]),Value_Matrix_Untransposed)

# Transpose the matrix

InterMediate = zip(*Value_Matrix_Untransposed)

Value_Matrix_0=[list(x) for x in InterMediate]




####### Check OK
# V= [column][row]

Vsum=[0 for i in range(num_of_columns)]


Cooperation_Group = [[0,1],4]
def NEW_Value(Cooperation_Group):
  columns_take = Take_Selected_Columns(Cooperation_Group)
  V_New_UN = []
  for i in range(len(Value_Matrix_Untransposed)):
    if i in columns_take:
      V_New_UN.append(Value_Matrix_Untransposed[i])
  Inter = zip(*V_New_UN)
  V_New = [list(x) for x in Inter]
  return V_New
Value_Matrix = NEW_Value(Cooperation_Group)



def Matrix_ixj(i,j):
  Matri = [[0 for col in range(j)] for row in range(i)]
  return Matri

Distribution_Matrix = Matrix_ixj(num_of_rows,num_of_columns)

Value_Distribution_Matrix = Matrix_ixj(num_of_rows,num_of_columns)

def find_indices(l, value):
    return [
        index for index, item in enumerate(l)
        if item == value
    ]

###### DATA loading
CorrelationWorkBook = load_workbook("/Users/Disk/Programs/IMMC2024winter/Code/tables/ItermCorrelation.xlsx")  # Work Book
Gain_WorkSheet = CorrelationWorkBook.get_sheet_by_name('Gain-Table')  # Work Sheet
Loss_WorkSheet = CorrelationWorkBook.get_sheet_by_name('Loss-Table')  # Work Sheet 

Correlation_Column_Names = list(string.ascii_uppercase[1:])
Correlation_Column_Names.append('AA')
Correlation_Column_Names.append('AB')
Correlation_Column_Names.append('AC')
Correlation_Column_Names.append('AD')
Correlation_Column_Names.append('AE')


Gain_Matrix = []
Loss_Matrix = []

for i in Correlation_Column_Names:
  column = Gain_WorkSheet[i]  # Column
  column_list = [column[x].value for x in range(1,1+V_size[1])] # take the number datas
  Gain_Matrix.append(column_list) # The V matrix, by name and item number

for i in Correlation_Column_Names:
  column = Loss_WorkSheet[i]  # Column
  column_list = [column[x].value for x in range(1,1+V_size[1])] # take the number datas
  Loss_Matrix.append(column_list) # The V matrix, by name and item number

Correlation_Coefficient = 10



def Update_Value_Matrix(item_distributed,who_to_give, Value_Matrix): #item_distributed is the location or index of the itgem in the order list
  V_Current = Value_Matrix 
  who_doesnot_get = [i for i in range(num_of_columns)]
  who_doesnot_get.pop(who_doesnot_get.index(who_to_give))
  # Change only after the cargo
  # Range_to_be_changed = range(item_distributed+1,len(Distribution_Order))
  # Change everything
  Range_to_be_changed = range(0,len(Distribution_Order)) 
  for changing_position in Range_to_be_changed: ##### If you wanna go from the beginning, it is really great
    item_order = Distribution_Order[item_distributed]
    changing_item_order = Distribution_Order[changing_position]
    # Q = Gain_Matrix[item_order-1][changing_item_order-1]
    # print(Q)
    R = (100+Correlation_Coefficient*Gain_Matrix[item_order][changing_item_order])/100*V_Current[changing_position][who_to_give] # Both the variables are item number not location 
    
    Value_Matrix[changing_position][who_to_give] = R
    for else_people in who_doesnot_get:
      R = (100+Correlation_Coefficient*Loss_Matrix[item_order][changing_item_order])/100*V_Current[changing_position][else_people] # Both the variables are item number not location n
      Value_Matrix[changing_position][else_people] = R
  return None
# This is right!!!

# Please notice that the numbering in the ItermCorrelation Matrix and the RawData Matrix is not the same.





def Calc_T_Expected(Value_Matrix):
  T_Expected = [] # Constant of the expected total value
  for row in range(num_of_rows):
    S = 0
    for col in range(num_of_columns):
      S += Value_Matrix[row][col]
    T_Expected.append(S)
  return T_Expected


def Justice_Value(col,T_Gained,T_Expected):
  r = T_Gained[col]/T_Expected[col]
  return r




J = [0 for i in range(num_of_columns)]


current_row_number = 0
T_Gained = [0 for i in range(num_of_columns)]
while current_row_number<num_of_rows:
  T_Expected = Calc_T_Expected(Value_Matrix)
  for col in range(num_of_columns):
    J[col]=Justice_Value(col,T_Gained,T_Expected)
  # Need a function that checks whether there're any repeated values in the list.
  whether_repeat_values = any(J.count(x) > 1 for x in J)
  
  # print(whether_repeat_values)
  
  if whether_repeat_values: 
    repeat_list = find_indices(J,0)
    Mediator = []

    for x in range(num_of_columns-1,-1,-1):
      if find_indices(J,0).count(x) == True:
        Mediator.append(x)     # POP out the not 0 ones

    V_mediator = []
    for x in Mediator:
      V_mediator.append(Value_Matrix[current_row_number][x])
    who_to_give = Value_Matrix[current_row_number].index(max(V_mediator))
  else:
    who_to_give = J.index(min(J)) # Anchor point. Via changing this, the optimization instance variable can be changed.
  Update_Value_Matrix(current_row_number,who_to_give,Value_Matrix)
  Distribution_Matrix[current_row_number][who_to_give] = 1 # Record the resultant matrix
  T_Gained[who_to_give]+=Value_Matrix[current_row_number][who_to_give]
  current_row_number += 1



def Calc_VD_Matrix(Value_Matrix,Distribution_Matrix): # D,V
  VD_Matrix = Matrix_ixj(num_of_rows,num_of_columns)  
  for rows in range(num_of_rows):
    for cols in range(num_of_columns):
      VD_Matrix[rows][cols] = Value_Matrix[rows][cols]*Distribution_Matrix[rows][cols]
  return VD_Matrix


Value_Distribution_Matrix = Calc_VD_Matrix(Value_Matrix,Distribution_Matrix)

print(Distribution_Matrix)

def Distribution_basedon_preference(group,Value_Matrix,Distribution_Matrix):
  cargo_list = []
  for row in range(num_of_rows):
    for i in group:
      if Distribution_Matrix[row][i] == 1:
        cargo_list.append(row)
  print(cargo_list)
  subVM = []
  for row in range(len(Value_Matrix)):
    subVM.append([])
    for col in range(len(Value_Matrix[row])):
      if col in group:
        subVM[row].append(Value_Matrix[row][col])

  Redistribution_matri = []
  for item in range(len(cargo_list)):
    Redistribution_matri.append([])
    if cargo_list[item]==1:
      who_to_give = subVM[item].index(max(subVM[item]))
      Redistribution_matri[item].append(who_to_give)

  return Redistribution_matri



  # Compare VM[row][i] and VM[row][]
    


r = Distribution_basedon_preference([0,1],Value_Matrix,Distribution_Matrix)
print(r)


#def Redistribution(Value_Matrix,Distribution_Matrix):
  # first pull the groups out from the cooperation_groups
    


# 计算完重新分配结果后根据分组情况把数据还原。